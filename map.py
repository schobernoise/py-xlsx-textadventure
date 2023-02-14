import utils
import openpyxl


class MapTile:
    def __init__(self, x, y, name):
        self.x = x
        self.y = y
        self.name = name

    def display_intro_text(self):
        raise NotImplementedError()

    def modify_player(self, player):
        raise NotImplementedError()


class MapBuilder:
    def __init__(self, excel_map) -> None:
        self.boundaries = (0, 0)
        self.world_map = self.map_from_xlsx(excel_map)

        self.return_map()

    def map_from_xlsx(self, excel_file):
        # TODO implement utils.get_workbook fn
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        world_map = []
        for x in range(2, sheet.max_row + 1):
            x_row = []
            for y in range(2, sheet.max_column + 1):
                x_row.append(MapTile(x, y, sheet.cell(x, y).value))
            world_map.append(x_row)
        self.boundaries = (sheet.max_row + 1, sheet.max_column + 1)
        return world_map

    def return_map(self):
        # for x, row in enumerate(self.world_map):
        #     for y, column in enumerate(row):
        #         print(self.world_map[x][y].name)
        return self.world_map

    def get_boundaries(self):
        return self.boundaries

    def return_tile(self, x, y):
        return self.world_map[x][y]
